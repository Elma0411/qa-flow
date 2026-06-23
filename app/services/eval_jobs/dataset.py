# 文件作用：加载、标准化和切分评测数据集样本。
# 关联说明：为 run.py 提供评测样本读取和标准化输入。

from __future__ import annotations

import csv
import io
import json
import os
from dataclasses import dataclass, field
from hashlib import sha1
from typing import Any, Dict, Iterable, List, Literal, Optional, Sequence, Tuple

from fastapi import UploadFile


InputFormat = Literal["auto", "csv", "xlsx", "jsonl", "json"]


@dataclass(frozen=True)
class DatasetPreview:
    detected_format: str
    columns: List[str]
    sample_rows: List[Dict[str, Any]]
    total_rows: Optional[int] = None
    warnings: List[str] = field(default_factory=list)


@dataclass(frozen=True)
class BatchDatasetPreview:
    files: List[Dict[str, Any]]
    schema_consistent: bool
    shared_columns: List[str]
    warnings: List[str] = field(default_factory=list)


def _read_upload_bytes(upload_file: UploadFile) -> bytes:
    try:
        data = upload_file.file.read()
        return data if isinstance(data, (bytes, bytearray)) else bytes(data)
    finally:
        try:
            upload_file.file.seek(0)
        except Exception:
            pass


def _decode_bytes(data: bytes, encoding: Optional[str]) -> str:
    if encoding:
        return data.decode(str(encoding))
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def detect_input_format(upload_file: UploadFile, input_format: str) -> str:
    fmt = str(input_format or "auto").strip().lower()
    if fmt in {"csv", "xlsx", "jsonl", "json"}:
        return fmt
    name = str(upload_file.filename or "").lower()
    ext = os.path.splitext(name)[1]
    if ext in {".csv"}:
        return "csv"
    if ext in {".xlsx", ".xlsm"}:
        return "xlsx"
    if ext in {".jsonl"}:
        return "jsonl"
    if ext in {".json"}:
        return "json"
    # Fallback: treat as jsonl if it looks line-delimited, otherwise json/csv later.
    return "auto"


def _coerce_optional_positive_int(value: Any, *, field_name: str) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        resolved = int(value)
    except Exception as exc:
        raise ValueError(f"{field_name} 必须是正整数或空") from exc
    if resolved < 1:
        raise ValueError(f"{field_name} 必须大于等于 1")
    return resolved


def parse_file_ranges_json(raw: Optional[str], *, files_count: int) -> Dict[int, Dict[str, Optional[int]]]:
    text = str(raw or "").strip()
    if not text:
        return {}
    try:
        payload = json.loads(text)
    except Exception as exc:
        raise ValueError("file_ranges_json 不是合法 JSON") from exc
    if not isinstance(payload, list):
        raise ValueError("file_ranges_json 必须是数组")

    parsed: Dict[int, Dict[str, Optional[int]]] = {}
    for idx, item in enumerate(payload):
        if not isinstance(item, dict):
            raise ValueError(f"file_ranges_json[{idx}] 必须是对象")
        try:
            file_index = int(item.get("file_index"))
        except Exception as exc:
            raise ValueError(f"file_ranges_json[{idx}].file_index 必须是整数") from exc
        if file_index < 0 or file_index >= max(0, int(files_count)):
            raise ValueError(f"file_ranges_json[{idx}].file_index 超出范围: {file_index}")
        if file_index in parsed:
            raise ValueError(f"file_ranges_json 中 file_index 重复: {file_index}")

        row_start = _coerce_optional_positive_int(
            item.get("row_start"),
            field_name=f"file_ranges_json[{idx}].row_start",
        )
        row_end = _coerce_optional_positive_int(
            item.get("row_end"),
            field_name=f"file_ranges_json[{idx}].row_end",
        )
        if row_start is not None and row_end is not None and row_start > row_end:
            raise ValueError(f"file_ranges_json[{idx}] 的 row_start 不能大于 row_end")
        parsed[file_index] = {"row_start": row_start, "row_end": row_end}
    return parsed


def resolve_row_range(
    *,
    total_rows: int,
    row_start: Optional[int],
    row_end: Optional[int],
) -> Dict[str, Any]:
    total = max(0, int(total_rows or 0))
    start_req = _coerce_optional_positive_int(row_start, field_name="row_start")
    end_req = _coerce_optional_positive_int(row_end, field_name="row_end")
    if start_req is not None and end_req is not None and start_req > end_req:
        raise ValueError("row_start 不能大于 row_end")

    if total <= 0:
        return {
            "requested": {"start": start_req, "end": end_req},
            "resolved": {"start": None, "end": None},
            "resolved_zero_based": {"start": None, "end": None},
            "selected_rows": 0,
        }

    start_1 = start_req if start_req is not None else 1
    end_1 = end_req if end_req is not None else total
    start_1 = min(max(1, start_1), total)
    end_1 = min(max(1, end_1), total)
    if start_1 > end_1:
        raise ValueError("row_start 不能大于 row_end")
    start_0 = start_1 - 1
    end_0 = end_1 - 1
    return {
        "requested": {"start": start_req, "end": end_req},
        "resolved": {"start": start_1, "end": end_1},
        "resolved_zero_based": {"start": start_0, "end": end_0},
        "selected_rows": max(0, end_0 - start_0 + 1),
    }


def _normalize_ws_for_hash(text: str) -> str:
    # Match unsupervised evaluation context-group normalization.
    norm = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    norm = " ".join(norm.split())
    return norm.strip()


def compute_group_id(context: str) -> str:
    norm = _normalize_ws_for_hash(context)
    return "sha1:" + sha1(norm.encode("utf-8")).hexdigest()


def compute_row_id(
    *,
    dataset_name: str,
    task_id: str,
    original_filename: str,
    question: str,
    answer: str,
    provided_id: Optional[str],
) -> str:
    provided = str(provided_id or "").strip()
    if provided:
        ds = str(dataset_name or "").strip()
        if ds:
            return f"human:{ds}:{provided}"
        return f"human::{provided}"
    payload = f"{task_id}{original_filename}{question}{answer}"
    return sha1(payload.encode("utf-8")).hexdigest()


def _parse_csv_rows(
    text: str,
    *,
    delimiter: str = ",",
) -> Tuple[List[str], List[Dict[str, Any]]]:
    delim = str(delimiter or ",")
    if len(delim) != 1:
        raise ValueError("delimiter 必须是单个字符")
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    fieldnames = [str(x) for x in (reader.fieldnames or []) if str(x).strip()]
    rows: List[Dict[str, Any]] = []
    for row in reader:
        if not isinstance(row, dict):
            continue
        rows.append({str(k): row.get(k) for k in row.keys()})
    return fieldnames, rows


def _resolve_sheet_name(
    sheet_names: Sequence[str],
    sheet_name: Optional[str],
) -> str:
    if not sheet_names:
        raise ValueError("xlsx 文件无可用 sheet")
    raw = str(sheet_name or "").strip()
    if not raw:
        return str(sheet_names[0])
    if raw.isdigit():
        idx = int(raw)
        if idx <= 0 or idx > len(sheet_names):
            raise ValueError(f"sheet_name 序号超出范围: {idx} (1..{len(sheet_names)})")
        return str(sheet_names[idx - 1])
    if raw not in sheet_names:
        raise ValueError(f"sheet_name 不存在: {raw}")
    return str(raw)


def _parse_xlsx_rows(
    data: bytes,
    *,
    sheet_name: Optional[str],
) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("缺少依赖 openpyxl，无法读取 .xlsx；请先 pip install openpyxl") from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    name = _resolve_sheet_name(wb.sheetnames, sheet_name)
    ws = wb[name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []

    columns: List[str] = []
    for cell in header:
        col = "" if cell is None else str(cell).strip()
        columns.append(col)

    fieldnames: List[str] = [c for c in columns if c]
    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        values = list(row)
        if not any(v is not None and str(v).strip() for v in values):
            continue
        item: Dict[str, Any] = {}
        for idx, col in enumerate(columns):
            if not col:
                continue
            item[col] = values[idx] if idx < len(values) else None
        rows.append(item)
    return fieldnames, rows


def _parse_jsonl_rows(text: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    cols: set = set()
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)
        if not isinstance(obj, dict):
            continue
        rows.append(obj)
        cols.update(obj.keys())
    return sorted(str(c) for c in cols), rows


def _coerce_json_to_rows(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if isinstance(payload, dict):
        if isinstance(payload.get("items"), list):
            return [x for x in payload.get("items") if isinstance(x, dict)]
        return [payload]
    return []


def _parse_json_rows(text: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    payload = json.loads(text)
    rows = _coerce_json_to_rows(payload)
    cols: set = set()
    for r in rows:
        cols.update(r.keys())
    return sorted(str(c) for c in cols), rows


def preview_dataset(
    upload_file: UploadFile,
    *,
    input_format: str = "auto",
    encoding: Optional[str] = None,
    delimiter: str = ",",
    sheet_name: Optional[str] = None,
    sample_size: int = 5,
) -> DatasetPreview:
    detected = detect_input_format(upload_file, input_format)
    warnings: List[str] = []

    data = _read_upload_bytes(upload_file)

    if detected == "xlsx":
        columns, rows = _parse_xlsx_rows(data, sheet_name=sheet_name)
        return DatasetPreview(
            detected_format="xlsx",
            columns=list(columns),
            sample_rows=rows[: max(1, int(sample_size or 5))],
            total_rows=len(rows),
            warnings=warnings,
        )

    text = _decode_bytes(data, encoding)

    if detected == "csv":
        columns, rows = _parse_csv_rows(text, delimiter=delimiter)
        return DatasetPreview(
            detected_format="csv",
            columns=list(columns),
            sample_rows=rows[: max(1, int(sample_size or 5))],
            total_rows=len(rows),
            warnings=warnings,
        )

    if detected == "jsonl":
        columns, rows = _parse_jsonl_rows(text)
        return DatasetPreview(
            detected_format="jsonl",
            columns=list(columns),
            sample_rows=rows[: max(1, int(sample_size or 5))],
            total_rows=len(rows),
            warnings=warnings,
        )

    if detected == "json":
        columns, rows = _parse_json_rows(text)
        return DatasetPreview(
            detected_format="json",
            columns=list(columns),
            sample_rows=rows[: max(1, int(sample_size or 5))],
            total_rows=len(rows),
            warnings=warnings,
        )

    # auto fallback: attempt jsonl -> json -> csv
    try:
        columns, rows = _parse_jsonl_rows(text)
        if rows:
            warnings.append("input_format=auto: detected as jsonl")
            return DatasetPreview(
                detected_format="jsonl",
                columns=list(columns),
                sample_rows=rows[: max(1, int(sample_size or 5))],
                total_rows=len(rows),
                warnings=warnings,
            )
    except Exception:
        pass
    try:
        columns, rows = _parse_json_rows(text)
        if rows:
            warnings.append("input_format=auto: detected as json")
            return DatasetPreview(
                detected_format="json",
                columns=list(columns),
                sample_rows=rows[: max(1, int(sample_size or 5))],
                total_rows=len(rows),
                warnings=warnings,
            )
    except Exception:
        pass
    try:
        columns, rows = _parse_csv_rows(text, delimiter=delimiter)
        warnings.append("input_format=auto: detected as csv")
        return DatasetPreview(
            detected_format="csv",
            columns=list(columns),
            sample_rows=rows[: max(1, int(sample_size or 5))],
            total_rows=len(rows),
            warnings=warnings,
        )
    except Exception as exc:
        raise ValueError(f"无法解析上传文件 (auto): {exc}") from exc


def preview_datasets(
    upload_files: Sequence[UploadFile],
    *,
    input_format: str = "auto",
    encoding: Optional[str] = None,
    delimiter: str = ",",
    sheet_name: Optional[str] = None,
    sample_size: int = 5,
    file_ranges_by_index: Optional[Dict[int, Dict[str, Optional[int]]]] = None,
    question_field: Optional[str] = None,
    answer_field: Optional[str] = None,
    context_field: Optional[str] = None,
    ref_answer_field: Optional[str] = None,
    id_field: Optional[str] = None,
    original_filename_field: Optional[str] = None,
) -> BatchDatasetPreview:
    files = list(upload_files or [])
    if not files:
        raise ValueError("至少需要上传一个文件")

    file_ranges = file_ranges_by_index or {}
    file_previews: List[Dict[str, Any]] = []
    warnings: List[str] = []
    baseline_columns: Optional[List[str]] = None
    baseline_column_set: Optional[set[str]] = None
    shared_column_set: Optional[set[str]] = None
    schema_consistent = True

    for file_index, upload_file in enumerate(files):
        preview = preview_dataset(
            upload_file,
            input_format=input_format,
            encoding=encoding,
            delimiter=delimiter,
            sheet_name=sheet_name,
            sample_size=sample_size,
        )
        columns = [str(col) for col in (preview.columns or [])]
        column_set = set(columns)
        if baseline_columns is None:
            baseline_columns = list(columns)
            baseline_column_set = set(column_set)
            shared_column_set = set(column_set)
        else:
            shared_column_set = set(shared_column_set or set()).intersection(column_set)

        baseline_set = set(baseline_column_set or set())
        missing_columns = sorted(baseline_set - column_set)
        extra_columns = sorted(column_set - baseline_set)
        schema_status = "ok" if not missing_columns and not extra_columns else "mismatch"
        if schema_status != "ok":
            schema_consistent = False

        row_cfg = file_ranges.get(file_index) or {}
        row_range = resolve_row_range(
            total_rows=int(preview.total_rows or 0),
            row_start=row_cfg.get("row_start"),
            row_end=row_cfg.get("row_end"),
        )

        file_previews.append(
            {
                "file_index": file_index,
                "filename": str(upload_file.filename or f"file_{file_index}"),
                "detected_format": preview.detected_format,
                "columns": columns,
                "sample_rows": preview.sample_rows,
                "total_rows": preview.total_rows,
                "warnings": preview.warnings,
                "mapping_check": mapping_check(
                    columns,
                    question_field=question_field,
                    answer_field=answer_field,
                    context_field=context_field,
                    ref_answer_field=ref_answer_field,
                    id_field=id_field,
                    original_filename_field=original_filename_field,
                ),
                "schema_status": schema_status,
                "missing_columns": missing_columns,
                "extra_columns": extra_columns,
                "row_range": row_range,
            }
        )
        warnings.extend([str(x) for x in preview.warnings])

    shared_columns: List[str] = []
    baseline_order = baseline_columns or []
    shared_set = set(shared_column_set or set())
    for column in baseline_order:
        if column in shared_set:
            shared_columns.append(column)

    return BatchDatasetPreview(
        files=file_previews,
        schema_consistent=schema_consistent,
        shared_columns=shared_columns,
        warnings=warnings,
    )


def load_dataset_rows(
    upload_file: UploadFile,
    *,
    input_format: str = "auto",
    encoding: Optional[str] = None,
    delimiter: str = ",",
    sheet_name: Optional[str] = None,
) -> Tuple[str, List[str], List[Dict[str, Any]]]:
    preview = preview_dataset(
        upload_file,
        input_format=input_format,
        encoding=encoding,
        delimiter=delimiter,
        sheet_name=sheet_name,
        sample_size=1,
    )
    detected = preview.detected_format
    data = _read_upload_bytes(upload_file)

    if detected == "xlsx":
        cols, rows = _parse_xlsx_rows(data, sheet_name=sheet_name)
        return detected, cols, rows

    text = _decode_bytes(data, encoding)
    if detected == "csv":
        cols, rows = _parse_csv_rows(text, delimiter=delimiter)
        return detected, cols, rows
    if detected == "jsonl":
        cols, rows = _parse_jsonl_rows(text)
        return detected, cols, rows
    if detected == "json":
        cols, rows = _parse_json_rows(text)
        return detected, cols, rows
    raise ValueError(f"unsupported detected_format: {detected}")


def load_dataset_rows_from_path(
    path: str,
    *,
    input_format: str = "auto",
    encoding: Optional[str] = None,
    delimiter: str = ",",
    sheet_name: Optional[str] = None,
) -> Tuple[str, List[str], List[Dict[str, Any]]]:
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"input file not found: {path}")
    filename = os.path.basename(path)
    with open(path, "rb") as f:
        data = f.read()

    fmt = str(input_format or "auto").strip().lower()
    if fmt not in {"auto", "csv", "xlsx", "jsonl", "json"}:
        fmt = "auto"

    if fmt == "auto":
        ext = os.path.splitext(filename.lower())[1]
        if ext == ".csv":
            fmt = "csv"
        elif ext in {".xlsx", ".xlsm"}:
            fmt = "xlsx"
        elif ext == ".jsonl":
            fmt = "jsonl"
        elif ext == ".json":
            fmt = "json"

    if fmt == "xlsx":
        cols, rows = _parse_xlsx_rows(data, sheet_name=sheet_name)
        return "xlsx", cols, rows

    text = _decode_bytes(data, encoding)
    if fmt == "csv":
        cols, rows = _parse_csv_rows(text, delimiter=delimiter)
        return "csv", cols, rows
    if fmt == "jsonl":
        cols, rows = _parse_jsonl_rows(text)
        return "jsonl", cols, rows
    if fmt == "json":
        cols, rows = _parse_json_rows(text)
        return "json", cols, rows

    # auto fallback
    try:
        cols, rows = _parse_jsonl_rows(text)
        if rows:
            return "jsonl", cols, rows
    except Exception:
        pass
    try:
        cols, rows = _parse_json_rows(text)
        if rows:
            return "json", cols, rows
    except Exception:
        pass
    cols, rows = _parse_csv_rows(text, delimiter=delimiter)
    return "csv", cols, rows


def extract_canonical_rows(
    raw_rows: Iterable[Dict[str, Any]],
    *,
    dataset_name: str,
    task_id: str,
    original_filename_default: str,
    question_field: str,
    answer_field: str,
    context_field: str,
    ref_answer_field: Optional[str] = None,
    id_field: Optional[str] = None,
    original_filename_field: Optional[str] = None,
) -> List[Dict[str, Any]]:
    qf = str(question_field or "").strip()
    af = str(answer_field or "").strip()
    cf = str(context_field or "").strip()
    if not (qf and af and cf):
        raise ValueError("question_field/answer_field/context_field 不能为空")

    rf = str(ref_answer_field or "").strip() or None
    idf = str(id_field or "").strip() or None
    fnf = str(original_filename_field or "").strip() or None

    items: List[Dict[str, Any]] = []
    for idx, row in enumerate(raw_rows):
        if not isinstance(row, dict):
            continue
        meta = dict(row)
        question = "" if row.get(qf) is None else str(row.get(qf))
        answer = "" if row.get(af) is None else str(row.get(af))
        context = "" if row.get(cf) is None else str(row.get(cf))
        ref_answer = None
        if rf:
            raw_ra = row.get(rf)
            if raw_ra is not None and str(raw_ra).strip() != "":
                ref_answer = str(raw_ra)

        original_filename = original_filename_default
        if fnf:
            raw_fn = row.get(fnf)
            if raw_fn is not None and str(raw_fn).strip():
                original_filename = str(raw_fn).strip()

        provided_id = None
        if idf:
            raw_id = row.get(idf)
            if raw_id is not None and str(raw_id).strip():
                provided_id = str(raw_id).strip()

        item_id = compute_row_id(
            dataset_name=dataset_name,
            task_id=task_id,
            original_filename=original_filename,
            question=question,
            answer=answer,
            provided_id=provided_id,
        )
        group_id = compute_group_id(context)
        items.append(
            {
                "id": item_id,
                "group_id": group_id,
                "task_id": task_id,
                "original_filename": original_filename,
                "context": context,
                "question": question,
                "answer": answer,
                "ref_answer": ref_answer,
                "meta": meta,
                "_row_index": idx,
            }
        )
    return items


def mapping_check(
    columns: Sequence[str],
    *,
    question_field: Optional[str] = None,
    answer_field: Optional[str] = None,
    context_field: Optional[str] = None,
    ref_answer_field: Optional[str] = None,
    id_field: Optional[str] = None,
    original_filename_field: Optional[str] = None,
) -> Dict[str, Any]:
    cols = {str(c) for c in (columns or [])}

    def _check(name: str, value: Optional[str], required: bool) -> Dict[str, Any]:
        val = str(value or "").strip()
        if not val:
            return {"required": required, "provided": False, "exists": False, "value": None}
        return {"required": required, "provided": True, "exists": val in cols, "value": val}

    return {
        "question_field": _check("question_field", question_field, True),
        "answer_field": _check("answer_field", answer_field, True),
        "context_field": _check("context_field", context_field, True),
        "ref_answer_field": _check("ref_answer_field", ref_answer_field, False),
        "id_field": _check("id_field", id_field, False),
        "original_filename_field": _check("original_filename_field", original_filename_field, False),
    }


__all__ = [
    "BatchDatasetPreview",
    "DatasetPreview",
    "compute_group_id",
    "compute_row_id",
    "detect_input_format",
    "extract_canonical_rows",
    "load_dataset_rows",
    "load_dataset_rows_from_path",
    "mapping_check",
    "parse_file_ranges_json",
    "preview_dataset",
    "preview_datasets",
    "resolve_row_range",
]
