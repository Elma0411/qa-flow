# 文件作用：将 CSV 数据集转换为 context/question/answer 三元组 JSONL。
# 关联说明：与其他 convert/filter 脚本并列，产出 baseline_eval 可消费的 triples。

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


@dataclass(frozen=True)
class ColumnSpec:
    context: str
    question: str
    answer: str


@dataclass(frozen=True)
class ColumnSelector:
    index_0: Optional[int] = None
    name: Optional[str] = None


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "将 CSV/Excel 的部分行转换为本项目 triples JSONL（dataset/split/id/context/question/answer）。\n"
            "支持通过 --rows 选择特定行，并通过 --*-col 指定列名（CSV）或列字母/序号（Excel）。"
        ),
    )
    parser.add_argument("--input", "-i", type=str, required=True, help="输入 CSV 或 Excel(.xlsx/.xlsm) 路径")
    parser.add_argument("--output", "-o", type=str, required=True, help="输出 triples JSONL 路径")
    parser.add_argument(
        "--sheet",
        type=str,
        default="",
        help="Excel 工作表（仅 .xlsx/.xlsm 生效）：可填名称或 1-based 序号；默认第 1 个 sheet",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help=(
            "Excel 表头所在行号（仅 .xlsx/.xlsm 生效，1-based；默认 1）。"
            "设为 0 表示无表头（此时可用列字母如 C/E/H 选择列）。"
        ),
    )
    parser.add_argument(
        "--rows",
        type=str,
        default="",
        help=(
            "选择要导出的行（数据行，去掉表头后从 1 开始）；支持如 '1,3-5,10'；"
            "空字符串表示导出全部行。"
        ),
    )
    parser.add_argument(
        "--encoding",
        type=str,
        default="utf-8-sig",
        help="CSV 文件编码（默认 utf-8-sig，可兼容带 BOM 的 UTF-8）",
    )
    parser.add_argument(
        "--delimiter",
        type=str,
        default=",",
        help="CSV 分隔符（默认 ','）",
    )

    parser.add_argument(
        "--context-col",
        type=str,
        default="",
        help="context 对应的列名（可选；为空时尝试自动推断）",
    )
    parser.add_argument(
        "--question-col",
        type=str,
        default="",
        help="question 对应的列名（可选；为空时尝试自动推断）",
    )
    parser.add_argument(
        "--answer-col",
        type=str,
        default="",
        help="answer 对应的列名（可选；为空时尝试自动推断）",
    )

    parser.add_argument(
        "--dataset",
        type=str,
        default="",
        help="写入 dataset 字段（默认从输入文件名自动生成，如 csv_input）",
    )
    parser.add_argument(
        "--split",
        type=str,
        default="custom",
        help="写入 split 字段（默认 custom）",
    )
    parser.add_argument(
        "--id-prefix",
        type=str,
        default="",
        help="id 前缀（默认自动生成：{dataset}_{split}）",
    )
    parser.add_argument(
        "--id-start",
        type=int,
        default=1,
        help="id 自增起始值（默认 1）",
    )
    parser.add_argument(
        "--id-width",
        type=int,
        default=0,
        help="id 序号零填充宽度（默认 0 不填充；如 6 → 000001）",
    )
    parser.add_argument(
        "--skip-empty",
        action="store_true",
        help="跳过 context/question/answer 任一为空的行（默认不跳过）",
    )
    return parser.parse_args()


def _auto_dataset_name(input_path: str) -> str:
    stem = os.path.splitext(os.path.basename(input_path))[0]
    token = re.sub(r"[^A-Za-z0-9]+", "_", stem).strip("_")
    if not token:
        token = "csv"
    if token.lower().startswith("csv_") or token.lower() == "csv":
        return token
    return f"csv_{token}"


def _parse_row_ranges(spec: str) -> List[Tuple[int, int]]:
    """
    Parse a row spec like: "1,3-5,10" into inclusive ranges (1-based).
    Empty string => [] meaning "all rows".
    """
    spec = (spec or "").strip()
    if not spec:
        return []

    ranges: List[Tuple[int, int]] = []
    for raw in spec.split(","):
        token = raw.strip()
        if not token:
            continue
        if re.fullmatch(r"\d+", token):
            n = int(token)
            if n <= 0:
                raise ValueError(f"--rows 中的行号必须 >= 1，收到: {token}")
            ranges.append((n, n))
            continue

        m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", token)
        if m:
            start = int(m.group(1))
            end = int(m.group(2))
            if start <= 0 or end <= 0:
                raise ValueError(f"--rows 中的行号必须 >= 1，收到: {token}")
            if end < start:
                raise ValueError(f"--rows 中范围必须 start<=end，收到: {token}")
            ranges.append((start, end))
            continue

        raise ValueError(f"无法解析 --rows 片段: {token!r}（示例：'1,3-5,10'）")

    ranges.sort()
    merged: List[Tuple[int, int]] = []
    for start, end in ranges:
        if not merged:
            merged.append((start, end))
            continue
        last_start, last_end = merged[-1]
        if start <= last_end + 1:
            merged[-1] = (last_start, max(last_end, end))
        else:
            merged.append((start, end))
    return merged


def _row_selected(row_no: int, ranges: Sequence[Tuple[int, int]]) -> bool:
    if not ranges:
        return True
    # ranges are sorted and merged
    lo = 0
    hi = len(ranges) - 1
    while lo <= hi:
        mid = (lo + hi) // 2
        start, end = ranges[mid]
        if row_no < start:
            hi = mid - 1
        elif row_no > end:
            lo = mid + 1
        else:
            return True
    return False


def _normalize_field(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def _pick_first_existing(fieldnames: Sequence[str], candidates: Sequence[str]) -> Optional[str]:
    existing = {f.strip(): f.strip() for f in fieldnames if isinstance(f, str) and f.strip()}
    for c in candidates:
        if c in existing:
            return c
    return None


def _resolve_columns(fieldnames: Sequence[str], args: argparse.Namespace) -> ColumnSpec:
    if not fieldnames:
        raise ValueError("CSV 未检测到表头，请确认文件格式或编码。")

    context_col = (args.context_col or "").strip()
    question_col = (args.question_col or "").strip()
    answer_col = (args.answer_col or "").strip()

    if not context_col:
        context_col = (
            _pick_first_existing(
                fieldnames,
                (
                    "context",
                    "content",
                    "text",
                    "passage",
                    "article",
                    "extracted_content",
                    "extracted_text",
                    "source_text",
                ),
            )
            or ""
        )
    if not question_col:
        question_col = _pick_first_existing(fieldnames, ("question", "q")) or ""
    if not answer_col:
        answer_col = _pick_first_existing(fieldnames, ("answer", "a")) or ""

    missing: List[str] = []
    for name, col in (("context", context_col), ("question", question_col), ("answer", answer_col)):
        if not col:
            missing.append(name)
            continue
        if col not in fieldnames:
            missing.append(f"{name}={col}")

    if missing:
        cols = ", ".join([c for c in fieldnames if isinstance(c, str)])
        raise ValueError(
            "无法确定 CSV 列映射（缺失/不存在）："
            + ", ".join(missing)
            + f"\n可用列名: {cols}\n"
            "请使用 --context-col/--question-col/--answer-col 显式指定。"
        )

    return ColumnSpec(context=context_col, question=question_col, answer=answer_col)


def _iter_csv_rows(
    path: str,
    encoding: str,
    delimiter: str,
) -> Tuple[Sequence[str], Iterable[Dict[str, Any]]]:
    """
    Return (fieldnames, rows_iter). Caller is responsible for iterating rows immediately.
    """
    f = open(path, "r", encoding=encoding, newline="")
    try:
        reader = csv.DictReader(f, delimiter=delimiter)
    except Exception:
        f.close()
        raise

    fieldnames = reader.fieldnames or []

    def _rows() -> Iterable[Dict[str, Any]]:
        try:
            for row in reader:
                if row is None:
                    continue
                yield row
        finally:
            f.close()

    return fieldnames, _rows()


def _is_excel(path: str) -> bool:
    ext = os.path.splitext(path)[1].lower()
    return ext in {".xlsx", ".xlsm"}


def _excel_letter_to_index_0(letter: str) -> int:
    """
    Convert Excel column letters to 0-based index: A->0, B->1, ..., Z->25, AA->26.
    """
    s = letter.strip().upper()
    if not s or not re.fullmatch(r"[A-Z]+", s):
        raise ValueError(f"非法 Excel 列字母: {letter!r}")

    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _index_0_to_excel_letter(index_0: int) -> str:
    if index_0 < 0:
        raise ValueError("index_0 must be >= 0")
    n = index_0 + 1
    out = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(ord("A") + rem) + out
    return out


def _resolve_selector_xlsx(
    fieldnames: Sequence[str],
    raw: str,
    role: str,
) -> ColumnSelector:
    """
    For Excel:
    - digits => 1-based column index
    - letters => Excel column letter(s), e.g. C / AA
    - otherwise => header name (exact match)
    """
    spec = (raw or "").strip()
    if not spec:
        return ColumnSelector(name=None, index_0=None)

    if re.fullmatch(r"\d+", spec):
        idx_1 = int(spec)
        if idx_1 <= 0:
            raise ValueError(f"{role} 列序号必须 >= 1，收到: {spec}")
        return ColumnSelector(index_0=idx_1 - 1, name=None)

    if re.fullmatch(r"[A-Za-z]+", spec):
        return ColumnSelector(index_0=_excel_letter_to_index_0(spec), name=None)

    if spec in fieldnames:
        return ColumnSelector(index_0=None, name=spec)

    cols = ", ".join([c for c in fieldnames if isinstance(c, str) and c.strip()])
    raise ValueError(f"Excel 未找到列名 {spec!r}（{role}）。可用列名: {cols}")


def _iter_xlsx_rows(
    path: str,
    sheet_spec: str,
    header_row: int,
) -> Tuple[Sequence[str], Iterable[Tuple[Any, ...]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("缺少依赖 openpyxl，无法读取 .xlsx 文件；请先 pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_spec and str(sheet_spec).strip():
            token = str(sheet_spec).strip()
            if re.fullmatch(r"\d+", token):
                idx_1 = int(token)
                if idx_1 <= 0 or idx_1 > len(wb.sheetnames):
                    raise ValueError(f"--sheet 超出范围: {idx_1}（共有 {len(wb.sheetnames)} 个 sheet）")
                ws = wb[wb.sheetnames[idx_1 - 1]]
            else:
                if token not in wb.sheetnames:
                    raise ValueError(f"--sheet 未找到: {token!r}（可用: {', '.join(wb.sheetnames)}）")
                ws = wb[token]
        else:
            ws = wb[wb.sheetnames[0]]

        max_col = int(ws.max_column or 0)
        if max_col <= 0:
            raise ValueError("Excel 为空或无法识别列数。")

        hdr = int(header_row)
        if hdr < 0:
            raise ValueError("--header-row 不能为负数。")

        if hdr > 0:
            header_values = next(
                ws.iter_rows(min_row=hdr, max_row=hdr, min_col=1, max_col=max_col, values_only=True)
            )
            fieldnames: List[str] = []
            seen: Dict[str, int] = {}
            for i, v in enumerate(header_values, start=1):
                name = _normalize_field(v)
                if not name:
                    name = _index_0_to_excel_letter(i - 1)
                if name in seen:
                    seen[name] += 1
                    name = f"{name}__{seen[name]}"
                else:
                    seen[name] = 0
                fieldnames.append(name)
            start_row = hdr + 1
        else:
            fieldnames = [_index_0_to_excel_letter(i) for i in range(max_col)]
            start_row = 1

        def _rows() -> Iterable[Tuple[Any, ...]]:
            try:
                for row in ws.iter_rows(min_row=start_row, min_col=1, max_col=max_col, values_only=True):
                    yield row
            finally:
                wb.close()

        return fieldnames, _rows()
    except Exception:
        wb.close()
        raise


def main() -> None:
    args = _parse_args()

    if not os.path.exists(args.input):
        raise SystemExit(f"输入文件不存在: {args.input}")

    row_ranges: List[Tuple[int, int]] = []
    try:
        row_ranges = _parse_row_ranges(str(args.rows))
    except ValueError as e:
        raise SystemExit(str(e)) from e

    dataset = (args.dataset or "").strip() or _auto_dataset_name(args.input)
    split = str(args.split or "custom").strip() or "custom"

    id_prefix = (args.id_prefix or "").strip() or f"{dataset}_{split}"
    id_start = max(1, int(args.id_start))
    id_width = max(0, int(args.id_width))

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)

    is_excel = _is_excel(args.input)
    if is_excel:
        try:
            header_row = int(args.header_row)
        except Exception as e:
            raise SystemExit(f"--header-row 不是整数: {args.header_row!r}") from e
        try:
            fieldnames, rows_iter = _iter_xlsx_rows(
                args.input,
                sheet_spec=str(args.sheet or ""),
                header_row=header_row,
            )
        except Exception as e:
            raise SystemExit(str(e)) from e

        try:
            # allow column letter / index / name in --*-col
            ctx_sel = _resolve_selector_xlsx(fieldnames, str(args.context_col or ""), role="context")
            q_sel = _resolve_selector_xlsx(fieldnames, str(args.question_col or ""), role="question")
            a_sel = _resolve_selector_xlsx(fieldnames, str(args.answer_col or ""), role="answer")
        except ValueError as e:
            raise SystemExit(str(e)) from e

        # If user didn't provide selectors, try name-based auto inference on header
        if ctx_sel.index_0 is None and ctx_sel.name is None:
            guess = _pick_first_existing(
                fieldnames,
                (
                    "context",
                    "content",
                    "text",
                    "passage",
                    "article",
                    "extracted_content",
                    "extracted_text",
                    "source_text",
                    "题目来源-相关段落\n（填写相关文件中对应段落或条款等）",
                    "题目来源-相关段落（填写相关文件中对应段落或条款等）",
                ),
            )
            if guess:
                ctx_sel = ColumnSelector(name=guess)
        if q_sel.index_0 is None and q_sel.name is None:
            guess = _pick_first_existing(fieldnames, ("question", "q", "题干\n（可参考评测数据集构建示例）", "题干（可参考评测数据集构建示例）"))
            if guess:
                q_sel = ColumnSelector(name=guess)
        if a_sel.index_0 is None and a_sel.name is None:
            guess = _pick_first_existing(fieldnames, ("answer", "a", "答案\n（可参考评测数据集构建示例）", "答案（可参考评测数据集构建示例）"))
            if guess:
                a_sel = ColumnSelector(name=guess)

        # Final check
        if ctx_sel.index_0 is None and ctx_sel.name is None:
            raise SystemExit("无法确定 context 列：请用 --context-col 指定（可用列字母如 H）。")
        if q_sel.index_0 is None and q_sel.name is None:
            raise SystemExit("无法确定 question 列：请用 --question-col 指定（可用列字母如 C）。")
        if a_sel.index_0 is None and a_sel.name is None:
            raise SystemExit("无法确定 answer 列：请用 --answer-col 指定（可用列字母如 E）。")
    else:
        fieldnames, rows_iter = _iter_csv_rows(
            args.input,
            encoding=str(args.encoding),
            delimiter=str(args.delimiter),
        )

        try:
            col_spec = _resolve_columns(fieldnames, args=args)
        except ValueError as e:
            raise SystemExit(str(e)) from e
        ctx_sel = ColumnSelector(name=col_spec.context)
        q_sel = ColumnSelector(name=col_spec.question)
        a_sel = ColumnSelector(name=col_spec.answer)

    total = 0
    selected = 0
    written = 0
    skipped_empty = 0

    with open(args.output, "w", encoding="utf-8", newline="\n") as out:
        for row_no, row in enumerate(rows_iter, start=1):
            total += 1
            if not _row_selected(row_no, row_ranges):
                continue

            selected += 1
            if isinstance(row, dict):
                context = _normalize_field(row.get(str(ctx_sel.name)))
                question = _normalize_field(row.get(str(q_sel.name)))
                answer = _normalize_field(row.get(str(a_sel.name)))
            else:
                values = list(row)

                def _get_value(sel: ColumnSelector) -> Any:
                    if sel.index_0 is not None:
                        return values[sel.index_0] if sel.index_0 < len(values) else None
                    if sel.name is None:
                        return None
                    try:
                        idx = list(fieldnames).index(sel.name)
                    except ValueError:
                        return None
                    return values[idx] if idx < len(values) else None

                context = _normalize_field(_get_value(ctx_sel))
                question = _normalize_field(_get_value(q_sel))
                answer = _normalize_field(_get_value(a_sel))

            if args.skip_empty and (not context or not question or not answer):
                skipped_empty += 1
                continue

            seq = id_start + (written)
            seq_str = str(seq).zfill(id_width) if id_width else str(seq)
            record = {
                "dataset": dataset,
                "split": split,
                "id": f"{id_prefix}_{seq_str}",
                "context": context,
                "question": question,
                "answer": answer,
            }
            out.write(json.dumps(record, ensure_ascii=False) + "\n")
            written += 1

    input_label = "Excel" if is_excel else "CSV"
    print(
        "转换完成:"
        f" {input_label} 共 {total} 行；匹配选择 {selected} 行；"
        f" 跳过空字段 {skipped_empty} 行；"
        f" 输出 {written} 行 -> {args.output}"
    )


if __name__ == "__main__":  # pragma: no cover
    main()
